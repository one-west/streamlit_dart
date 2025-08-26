import streamlit as st
import pandas as pd
import numpy as np
import re
from dotenv import load_dotenv
import OpenDartReader
import os
from datetime import datetime

# =========================
#  ë¬¸ìì—´ -> ìˆ«ì ë³€í™˜ (ê°•í™”)
# =========================
def to_number_strict(x):
    if pd.isna(x):
        return np.nan
    s = str(x)

    # ê³µë°±/ì œë¡œí­/ì†Œí”„íŠ¸í•˜ì´í”ˆ ì œê±°
    for ch in [
        "\u00a0", "\ufeff", "\u202f", "\u2009", "\u200a", "\u2007",
        "\u200b", "\u200c", "\u200d", "\u2060", "\u00ad"
    ]:
        s = s.replace(ch, "")

    # í†µí™”/ì²œë‹¨ìœ„ ì œê±°
    s = s.replace(",", "").replace("â‚©", "").replace("ì›", "").strip()

    # í•˜ì´í”ˆ/ë§ˆì´ë„ˆìŠ¤ í†µì¼
    s = (s.replace("\u2011", "-")   # non-breaking hyphen
           .replace("\u2212", "-")  # unicode minus
           .replace("â€“", "-")       # en dash
           .replace("â€”", "-"))      # em dash

    # ì‚¼ê°í˜• ìŒìˆ˜í‘œê¸° (â–³/â–²) â†’ ìŒìˆ˜
    s = re.sub(r"^[\u25B3\u25B2]\s*", "-", s)

    # ê´„í˜¸ ìŒìˆ˜í‘œê¸°: (1234) â†’ -1234
    if re.fullmatch(r"\(.*\)", s):
        s = "-" + s[1:-1].strip()

    # ì•ì˜ + ì œê±°
    if s.startswith("+"):
        s = s[1:]

    # ìˆ«ì/ë¶€í˜¸/ì†Œìˆ˜ì  ì™¸ ì œê±°
    s = re.sub(r"[^0-9\-\.+]", "", s)

    if s in ("", "-", "--", "+"):
        return np.nan

    return pd.to_numeric(s, errors="coerce")


# =========================
#  ì—‘ì…€ ì €ì¥ (XlsxWriter + openpyxl 2ì°¨ êµì •)
# =========================
def save_excel_with_comma_format(df: pd.DataFrame, file_name: str):
    """
    1) ëª¨ë“  '*amount' ì—´ì„ ìˆ«ìí˜•ìœ¼ë¡œ ë³´ì •
    2) XlsxWriterë¡œ 'ì…€ ë‹¨ìœ„' ì‘ì„± (ìˆ«ìâ†’write_number, ê²°ì¸¡â†’write_blank) + #,##0 ì„œì‹
    3) openpyxlë¡œ 2ì°¨ ê²€ì‚¬/êµì •: í˜¹ì‹œ ë‚¨ì€ ë¬¸ìì—´ ì…€ì€ ìˆ«ìë¡œ ê°•ì œ ë³€í™˜ + #,##0 ì„œì‹
    """
    import math, re
    from openpyxl import load_workbook

    def _norm_header(h: str) -> str:
        if h is None:
            return ""
        t = str(h).lower()
        for ch in ["\u00a0", "\ufeff", "\u202f", "\u2009", "\u200a", "\u2007", "\u200b", "\u200c", "\u200d", "\u2060", "\u00ad"]:
            t = t.replace(ch, "")
        return t.strip()

    amount_cols = [c for c in df.columns if "amount" in _norm_header(c)]

    # 1) DF ìˆ«ìí˜• ë³´ì •
    for col in amount_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # 2) XlsxWriterë¡œ ì…€ ë‹¨ìœ„ ì‘ì„±
    with pd.ExcelWriter(file_name, engine="xlsxwriter") as writer:
        wb = writer.book
        ws = wb.add_worksheet("Sheet1")
        writer.sheets["Sheet1"] = ws

        fmt_num   = wb.add_format({"num_format": "#,##0"})
        fmt_text  = wb.add_format()
        fmt_blank = wb.add_format({"num_format": "#,##0"})

        # í—¤ë”
        for j, col in enumerate(df.columns):
            ws.write(0, j, col, fmt_text)

        # ë°ì´í„°
        n_rows, n_cols = df.shape
        for i in range(n_rows):
            row = df.iloc[i]
            for j, col in enumerate(df.columns):
                val = row[col]
                if col in amount_cols:
                    if pd.isna(val):
                        ws.write_blank(i+1, j, None, fmt_blank)
                    else:
                        ws.write_number(i+1, j, float(val), fmt_num)
                else:
                    if pd.isna(val):
                        ws.write_blank(i+1, j, None)
                    elif isinstance(val, (int, float)) and not isinstance(val, bool) and math.isfinite(val):
                        ws.write_number(i+1, j, float(val))
                    else:
                        ws.write(i+1, j, str(val))

        ws.autofilter(0, 0, n_rows, n_cols-1)
        for j, col in enumerate(df.columns):
            if col in amount_cols:
                ws.set_column(j, j, 18, fmt_num)
            else:
                ws.set_column(j, j, 18)

    # 3) openpyxlë¡œ 2ì°¨ êµì •
    wb2 = load_workbook(file_name)
    ws2 = wb2.active

    # í—¤ë” ì¸ë±ìŠ¤ ë§¤í•‘
    header = [c.value for c in ws2[1]]
    name_to_colidx = {str(h): idx for idx, h in enumerate(header, start=1) if h is not None}

    def _to_number_strict_openpyxl(x):
        if x is None:
            return None
        s = str(x)
        for ch in ["\u00a0","\ufeff","\u202f","\u2009","\u200a","\u2007","\u200b","\u200c","\u200d","\u2060","\u00ad"]:
            s = s.replace(ch, "")
        s = (s.replace(",", "").replace("â‚©","").replace("ì›","").strip())
        s = (s.replace("\u2011","-").replace("\u2212","-").replace("â€“","-").replace("â€”","-"))
        s = re.sub(r"^[\u25B3\u25B2]\s*", "-", s)
        if re.fullmatch(r"\(.*\)", s):
            s = "-" + s[1:-1].strip()
        if s.startswith("+"):
            s = s[1:]
        s = re.sub(r"[^0-9\-\.+]", "", s)
        if s in ("", "-", "--", "+"):
            return None
        try:
            return float(s)
        except Exception:
            return None

    for col in amount_cols:
        col_idx = name_to_colidx.get(col)
        if not col_idx:
            # normalize ì´ë¦„ìœ¼ë¡œ ì¬íƒìƒ‰
            for h, idx in name_to_colidx.items():
                if "amount" in _norm_header(h) and _norm_header(h) == _norm_header(col):
                    col_idx = idx
                    break
        if not col_idx:
            continue

        for r in range(2, ws2.max_row + 1):
            cell = ws2.cell(row=r, column=col_idx)
            if isinstance(cell.value, str):
                num = _to_number_strict_openpyxl(cell.value)
                if num is not None:
                    cell.value = num
                    cell.number_format = "#,##0"
            else:
                if cell.value is not None:
                    cell.number_format = "#,##0"

    wb2.save(file_name)


# =========================
#  ì•± ë³¸ë¬¸
# =========================
load_dotenv()
api_key = os.getenv("DART_API_KEY") or st.secrets.get("DART_API_KEY", None)
if not api_key:
    api_key = st.sidebar.text_input("API í‚¤ë¥¼ ì…ë ¥í•˜ì„¸ìš”", type="password")
if not api_key:
    st.warning("API í‚¤ê°€ í•„ìš”í•©ë‹ˆë‹¤. ì…ë ¥ í›„ ë‹¤ì‹œ ì‹œë„í•˜ì„¸ìš”.")
    st.stop()

dart = OpenDartReader(api_key)

st.title("ğŸ“Š DART ì¬ë¬´ì œí‘œ ìˆ˜ì§‘ê¸°")
st.markdown("ì¢…ëª©ì½”ë“œë¥¼ ì…ë ¥í•˜ë©´ ì¬ë¬´ì œí‘œë¥¼ ê°€ì ¸ì˜µë‹ˆë‹¤.")

# ì¢…ëª©ì½”ë“œ â†’ ê¸°ì—…ëª… ë§¤í•‘
code_name_map = {
    "006400": "ì‚¼ì„±SDI",
    "373220": "LGì—ë„ˆì§€ì†”ë£¨ì…˜",
    "01592447": "ì—ìŠ¤ì¼€ì´ì˜¨",
    "259630": "ì— í”ŒëŸ¬ìŠ¤",
    "137400": "í”¼ì—”í‹°",
    "222080": "ì”¨ì•„ì´ì—ìŠ¤",
    "267320": "ë‚˜ì¸í…Œí¬",
    "196490": "ë””ì—ì´í…Œí¬ë†€ë¡œì§€",
    "109740": "ë””ì—ìŠ¤ì¼€ì´",
    "299030": "í•˜ë‚˜ê¸°ìˆ ",
    "240600": "ìœ ì§„í…Œí¬ë†€ë¡œì§€",
    "148930": "ì—ì´ì¹˜ì™€ì´í‹°ì”¨",
}
company_names = list(code_name_map.values())

# ì˜µì…˜: ì§„ë‹¨ ëª¨ë“œ
diag_mode = st.sidebar.toggle("ğŸ” ì§„ë‹¨ ëª¨ë“œ(ì €ì¥ ì „/í›„ thstrm_add_amount ì ê²€)", value=False)

select_all = st.checkbox("âœ… ì „ì²´ ì„ íƒ", value=True)
selected_names = st.multiselect(
    "ì¡°íšŒí•  ê¸°ì—… ì„ íƒ",
    options=company_names,
    default=company_names if select_all else [],
    key="corp_selector",
)
codes = [code for code, name in code_name_map.items() if name in selected_names]

current_year = datetime.now().year
year_range = list(range(current_year, current_year - 10, -1))
years = st.multiselect("ì¡°íšŒ ì—°ë„ (ë³µìˆ˜ ì„ íƒ ê°€ëŠ¥)", year_range, default=[current_year - 1])

report_map = {
    "ì‚¬ì—…ë³´ê³ ì„œ": "11011",
    "ë°˜ê¸°ë³´ê³ ì„œ": "11012",
    "3ë¶„ê¸°ë³´ê³ ì„œ": "11014",
    "1ë¶„ê¸°ë³´ê³ ì„œ": "11013",
}
report_label = st.selectbox("ë³´ê³ ì„œ ìœ í˜•", list(report_map.keys()))
report_code = report_map[report_label]

if st.button("ğŸ“¥ ì¬ë¬´ì œí‘œ ìˆ˜ì§‘"):
    if not codes:
        st.info("ì„ íƒëœ ê¸°ì—…ì´ ì—†ìŠµë‹ˆë‹¤.")
        st.stop()
    if not years:
        st.info("ì„ íƒëœ ì—°ë„ê°€ ì—†ìŠµë‹ˆë‹¤.")
        st.stop()

    result_list = []
    for year in years:
        for code in codes:
            try:
                df = dart.finstate_all(code, bsns_year=year, reprt_code=report_code)
                if isinstance(df, pd.DataFrame) and not df.empty:
                    df["ì¡°íšŒê¸°ì—…"] = code_name_map.get(code, code)
                    df["ì¡°íšŒì—°ë„"] = year
                    result_list.append(df)
                    st.success(f"{year} - {code} ìˆ˜ì§‘ ì™„ë£Œ")
                else:
                    st.warning(f"{year} - {code} ë°ì´í„° ì—†ìŒ")
            except Exception as e:
                st.error(f"{year} - {code} ì˜¤ë¥˜: {e}")

    if not result_list:
        st.info("ìˆ˜ì§‘ëœ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        st.stop()

    result_df = pd.concat(result_list, ignore_index=True)

    # ëª¨ë“  '*amount' ì—´ ìˆ«ìí™”
    amount_like_cols = [c for c in result_df.columns if "amount" in str(c).lower()]
    for col in amount_like_cols:
        result_df[col] = result_df[col].apply(to_number_strict)
        result_df[col] = pd.to_numeric(result_df[col], errors="coerce").astype("float64")

    # ì§„ë‹¨ ëª¨ë“œ: ì €ì¥ ì „ thstrm_add_amount ìƒíƒœ í™•ì¸
    if diag_mode and "thstrm_add_amount" in result_df.columns:
        target = "thstrm_add_amount"
        str_mask = result_df[target].map(lambda v: isinstance(v, str))
        st.info(f"[ì €ì¥ ì „] {target} dtype: {result_df[target].dtype}, ë¬¸ìì—´ê°œìˆ˜: {int(str_mask.sum())}, NaN: {int(result_df[target].isna().sum())}")

    # ì €ì¥
    file_name = f"dart_finstate_{'_'.join(map(str, years))}.xlsx"
    save_excel_with_comma_format(result_df, file_name)

    # ì§„ë‹¨ ëª¨ë“œ: ì €ì¥ í›„ ì‹¤ì œ ì…€ íƒ€ì… ê²€ì¦
    if diag_mode and "thstrm_add_amount" in result_df.columns:
        from openpyxl import load_workbook
        wb = load_workbook(file_name, data_only=True)
        ws = wb.active

        col_idx = None
        for c in ws[1]:
            if str(c.value).strip().lower() == "thstrm_add_amount":
                col_idx = c.col_idx
                break

        if col_idx is None:
            st.error("ì—‘ì…€ì—ì„œ 'thstrm_add_amount' ì—´ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤.")
        else:
            types = []
            bad_rows = []
            max_check = min(500, ws.max_row)
            for r in range(2, max_check + 1):
                cell = ws.cell(row=r, column=col_idx)
                types.append(cell.data_type)  # 'n'ì´ë©´ ìˆ«ì, 's'/'inlineStr'ë©´ í…ìŠ¤íŠ¸
                if cell.data_type not in (None, 'n') and cell.value is not None:
                    bad_rows.append((r, repr(cell.value), cell.number_format))

            st.info(f"[ì €ì¥ í›„] ê²€ì‚¬í–‰ìˆ˜: {max_check-1}, ìˆ«ìí˜•ì…€: {sum(t=='n' for t in types)}, ë¬¸ìí˜•ì…€: {sum((t not in (None,'n')) for t in types)}")
            if bad_rows:
                st.warning("ë¬¸ìí˜•ìœ¼ë¡œ ë‚¨ì€ ì…€(ì¼ë¶€):")
                st.write(bad_rows[:20])
            else:
                st.success("ëª¨ë“  ê²€ì‚¬ ì…€ì—ì„œ ìˆ«ìí˜•ìœ¼ë¡œ ì €ì¥ë˜ì—ˆìŠµë‹ˆë‹¤. (data_type='n')")

    with open(file_name, "rb") as f:
        st.download_button(
            label="ğŸ“ ì—‘ì…€ ë‹¤ìš´ë¡œë“œ",
            data=f,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.success("ì™„ë£Œë˜ì—ˆìŠµë‹ˆë‹¤! (ëª¨ë“  *amount ì—´ ìˆ«ìí˜• + ì—‘ì…€ #,##0 í¬ë§·)")
